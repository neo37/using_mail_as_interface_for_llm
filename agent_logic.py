import os
import json
import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from dotenv import load_dotenv
from openai import OpenAI

load_dotenv()

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

class SupplierLLMAgent:
    """
    LLM-агент для извлечения данных о товаре и генерации уточняющих вопросов.
    Теперь работает с моделью "gpt-4o-mini".
    """
    def __init__(self, required_fields=None):
        if required_fields is None:
            self.required_fields = ["product_name", "price", "dimensions", "weight", "material"]
        else:
            self.required_fields = required_fields

    def parse_supplier_answer(self, supplier_text: str) -> dict:
        """
        Отправляет текст поставщика в OpenAI и возвращает JSON-структуру с нужными полями.
        Если поля отсутствуют, оставляет пустые строки.
        """
        system_prompt = (
            "Ты — помощник, который анализирует ответ поставщика. "
            "Нужно извлечь ключевые поля товара и вернуть JSON-структуру. "
            "Если данные отсутствуют, оставь пустую строку."
        )
        user_prompt = (
            f"Поля, которые нужны: {', '.join(self.required_fields)}.\n"
            f"Вот ответ поставщика:\n---\n{supplier_text}\n---\n"
            "Верни результат ТОЛЬКО в формате JSON. "
            "Пример: {{\"product_name\": \"...\", \"price\": \"...\", ...}}"
        )

        try:
            response = client.chat.completions.create(
                model="gpt-4o-mini",  #
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=0.0
            )

            content = response.choices[0].message.content
            data = json.loads(content)
        except Exception as e:
            print(f"Ошибка при парсинге ответа LLM: {e}")
            data = {}

        clean_data = {}
        for field in self.required_fields:
            clean_data[field] = data.get(field, "")
        return clean_data


    def is_data_complete(self, data: dict) -> bool:
        """
        Проверяет, что все нужные поля заполнены.
        """
        return all(data.get(field) for field in self.required_fields)

    def generate_clarification_question(self, data: dict) -> str:
        """
        Если каких-то данных не хватает, генерирует уточняющий вопрос.
        """
        missing_fields = [f for f in self.required_fields if not data.get(f)]
        if not missing_fields:
            return ""

        system_prompt = (
            "Ты — человек, который общается с поставщиком. "
            "Тебе не хватает части данных. Напиши вежливый, но конкретный вопрос, "
            "чтобы попросить недостающие детали."
        )
        user_prompt = (
            "Мне не хватает данных о следующих полях: "
            f"{', '.join(missing_fields)}. "
            "Сформулируй короткий вежливый запрос, чтобы получить эти детали."
        )

        try:
            response = client.chat.completions.create(
                model="gpt-4o-mini",  # Используем ту же модель для генерации вопроса
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=0.7
            )
            # Доступ к содержимому через атрибуты
            question = response.choices[0].message.content
        except Exception as e:
            print(f"Ошибка при генерации уточняющего вопроса: {e}")
            question = "Пожалуйста, уточните недостающие данные."
        return question


class SupplierDataManager:
    """
    Хранит данные по каждому поставщику (ключ — email поставщика).
    """
    def __init__(self):
        self.data = {}

    def update_data(self, sender_email: str, new_fields: dict):
        if sender_email not in self.data:
            self.data[sender_email] = {}
        for k, v in new_fields.items():
            if v:
                self.data[sender_email][k] = v

    def is_complete(self, sender_email: str, required_fields: list) -> bool:
        stored = self.data.get(sender_email, {})
        return all(stored.get(f) for f in required_fields)

    def get_data(self, sender_email: str) -> dict:
        return self.data.get(sender_email, {})


class YandexEmailSender:
    """
    Отправка писем через SMTP (Яндекс).
    """
    def __init__(self):
        self.smtp_server = os.getenv("SMTP_SERVER", "smtp.yandex.ru")
        self.smtp_port = int(os.getenv("SMTP_PORT", "587"))
        self.sender_email = os.getenv("YANDEX_EMAIL")
        self.sender_password = os.getenv("YANDEX_PASSWORD")

    def reply_to_sender(self, recipient_email: str, subject: str, body: str):
        msg = MIMEMultipart()
        msg['From'] = self.sender_email
        msg['To'] = recipient_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        server = None
        try:
            server = smtplib.SMTP(self.smtp_server, self.smtp_port)
            server.starttls()
            server.login(self.sender_email, self.sender_password)
            server.send_message(msg)
            print(f"Отправлено письмо на {recipient_email} с темой '{subject}'")
        except Exception as e:
            print(f"Ошибка при отправке письма: {e}")
        finally:
            if server:
                server.quit()


def save_supplier_data_to_excel(data: dict, filename="suppliers_data.xlsx"):
    """
    Сохраняет данные вида:
      { "supplier_email_1": {"product_name": "...", ...}, ... }
    в Excel, где каждая строка — один поставщик.
    """
    # Собираем все поля, присутствующие в данных
    all_fields = set()
    for d in data.values():
        all_fields.update(d.keys())

    all_fields = list(all_fields)
    if not all_fields:
        print("Нет данных для сохранения.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suppliers"

    headers = ["supplier_email"] + all_fields
    for col_index, field in enumerate(headers, start=1):
        ws.cell(row=1, column=col_index, value=field)

    row_index = 2
    for supplier_email, fields_dict in data.items():
        ws.cell(row=row_index, column=1, value=supplier_email)
        for col_index, field in enumerate(all_fields, start=2):
            ws.cell(row=row_index, column=col_index, value=fields_dict.get(field, ""))
        row_index += 1

    wb.save(filename)
    print(f"Данные сохранены в {filename}")
