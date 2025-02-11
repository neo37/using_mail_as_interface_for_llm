import imaplib
import email
import time
import smtplib
import os
import openpyxl
import json
from dataclasses import dataclass
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT
# ============ LLM Агент для обработки данных о товаре ============
# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT
# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT
# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT# DRAFT
import openai
openai.api_key = os.getenv("OPENAI_API_KEY")  # Убедитесь, что в системе задан OPENAI_API_KEY

class SupplierLLMAgent:
    """
    Упрощённый LLM-агент, который пытается извлечь структуру данных из ответа поставщика
    и при необходимости генерирует уточняющий вопрос.
    """

    def __init__(self, required_fields=None):
        if required_fields is None:
            self.required_fields = ["product_name", "price", "dimensions", "weight", "material"]
        else:
            self.required_fields = required_fields

    def parse_supplier_answer(self, supplier_text: str) -> dict:
        """
        Запрашиваем у LLM структуру JSON на базе текста ответа.
        Если чего-то не хватает, поля будут пустыми.
        """
        system_prompt = (
            "Ты — помощник, который анализирует ответ поставщика. "
            "Нужно извлечь ключевые поля товара и вернуть JSON-структуру. "
            "Если данные отсутствуют, оставь пустую строку."
        )
        user_prompt = (
            f"Поля, которые нужны: {', '.join(self.required_fields)}.\n"
            f"Вот ответ поставщика:\n---\n{supplier_text}\n---\n"
            "Верни результат ТОЛЬКО в формате JSON. "
            "Пример: {\"product_name\": \"...\", \"price\": \"...\", ... }"
        )

        try:
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=0.0
            )
            content = response["choices"][0]["message"]["content"]

            # Пытаемся распарсить ответ как JSON
            data = json.loads(content)
        except Exception as e:
            print(f"Ошибка при парсинге ответа LLM: {e}")
            data = {}

        # На всякий случай "причесываем" результат
        clean_data = {}
        for field in self.required_fields:
            clean_data[field] = data.get(field, "")

        return clean_data

    def is_data_complete(self, data: dict) -> bool:
        """
        Проверяем, все ли нужные поля заполнены.
        """
        return all(data.get(field) for field in self.required_fields)

    def generate_clarification_question(self, data: dict) -> str:
        """
        Формируем уточняющий вопрос, если чего-то не хватает.
        """
        missing_fields = [f for f in self.required_fields if not data.get(f)]
        if not missing_fields:
            return ""

        system_prompt = (
            "Ты — человек, который общается с поставщиком. "
            "Тебе не хватает части данных. Напиши вежливый, но конкретный вопрос, "
            "чтобы попросить недостающие детали."
        )
        user_prompt = (
            "Мне не хватает данных о следующих полях: "
            f"{', '.join(missing_fields)}. "
            "Сформулируй короткий вежливый запрос, чтобы получить эти детали."
        )

        try:
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=0.7
            )
            question = response["choices"][0]["message"]["content"]
        except Exception as e:
            print(f"Ошибка при генерации уточняющего вопроса: {e}")
            question = "Пожалуйста, уточните недостающие данные."
        return question


# ============ Класс для хранения и обновления данных о поставщике ============

class SupplierDataManager:
    """
    Хранит собранные данные о поставщике в словаре.
    В реальном проекте стоит привязываться к конкретному треду/ID.
    """
    def __init__(self):
        self.data = {}  # ключ: email отправителя, значение: dict с полями

    def update_data(self, sender_email: str, new_fields: dict):
        """
        Объединяет старые и новые поля для конкретного поставщика (по email).
        """
        if sender_email not in self.data:
            self.data[sender_email] = {}
        for k, v in new_fields.items():
            if v:  # если поле не пустое, перезаписываем
                self.data[sender_email][k] = v

    def is_complete(self, sender_email: str, required_fields: list) -> bool:
        """
        Проверяем, заполнены ли все необходимые поля для данного поставщика.
        """
        stored = self.data.get(sender_email, {})
        return all(stored.get(f) for f in required_fields)

    def get_data(self, sender_email: str) -> dict:
        return self.data.get(sender_email, {})


# ============ Класс для приёма писем через IMAP Yandex ============

class YandexEmailReceiver:
    def __init__(self, imap_server: str, username: str, password: str, mailbox: str = 'inbox'):
        self.imap_server = imap_server
        self.username = username
        self.password = password
        self.mailbox = mailbox

    def fetch_unseen_emails(self):
        """
        Возвращает список (email_message, from_address) для всех непрочитанных писем.
        """
        results = []
        try:
            mail = imaplib.IMAP4_SSL(self.imap_server)
            mail.login(self.username, self.password)
            mail.select(self.mailbox)

            result, data = mail.search(None, 'UNSEEN')
            if result != 'OK':
                print("Ошибка при поиске непрочитанных писем.")
                return results

            email_ids = data[0].split()
            for e_id in email_ids:
                result, msg_data = mail.fetch(e_id, '(RFC822)')
                if result != 'OK':
                    print(f"Ошибка при получении письма id {e_id.decode()}")
                    continue

                raw_email = msg_data[0][1]
                msg = email.message_from_bytes(raw_email)

                # Парсим адрес отправителя
                from_addr = msg.get("From", "unknown")
                # Можно тут распарсить "From" детальнее, но упрощаем
                results.append((msg, from_addr))

            mail.logout()
        except Exception as e:
            print(f"Ошибка при работе с IMAP: {e}")

        return results


# ============ Класс для отправки писем через SMTP (Yandex) ============

class YandexEmailSender:
    def __init__(self, smtp_server: str, smtp_port: int, sender_email: str, sender_password: str):
        self.smtp_server = smtp_server
        self.smtp_port = smtp_port
        self.sender_email = sender_email
        self.sender_password = sender_password

    def reply_to_sender(self, recipient_email: str, subject: str, body: str):
        """
        Отправляем письмо на указанный адрес.
        """
        msg = MIMEMultipart()
        msg['From'] = self.sender_email
        msg['To'] = recipient_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        server = None
        try:
            server = smtplib.SMTP(self.smtp_server, self.smtp_port)
            server.starttls()
            server.login(self.sender_email, self.sender_password)
            server.send_message(msg)
            print(f"Отправлено письмо на {recipient_email} с темой '{subject}'")
        except Exception as e:
            print(f"Ошибка при отправке письма: {e}")
        finally:
            if server:
                server.quit()


# ============ Сохранение собранных данных в Excel ============

def save_supplier_data_to_excel(data: dict, filename="suppliers_data.xlsx"):
    """
    data: словарь вида {
      "supplier_email_1": {"product_name": "...", "price": "...", ...},
      "supplier_email_2": {...},
      ...
    }
    Сохраняем построчно, каждый поставщик — своя строка.
    """
    # Соберём все поля, какие вообще могут быть (допустим, как union)
    all_fields = set()
    for d in data.values():
        all_fields.update(d.keys())

    all_fields = list(all_fields)
    if not all_fields:
        print("Нет данных для сохранения")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suppliers"

    # Заголовки
    for col_index, field in enumerate(["supplier_email"] + all_fields, start=1):
        ws.cell(row=1, column=col_index, value=field)

    row_index = 2
    for supplier_email, fields_dict in data.items():
        ws.cell(row=row_index, column=1, value=supplier_email)
        for col_index, field in enumerate(all_fields, start=2):
            ws.cell(row=row_index, column=col_index, value=fields_dict.get(field, ""))
        row_index += 1

    wb.save(filename)
    print(f"Данные сохранены в {filename}")


# ============ Пример использования всего вместе ============

if __name__ == "__main__":

    # --- Настройки для Yandex IMAP/SMTP
    IMAP_SERVER = "imap.yandex.ru"
    SMTP_SERVER = "smtp.yandex.ru"
    SMTP_PORT = 587

    # Укажите свои реальные данные:
    YANDEX_EMAIL = "your_yandex_email@yandex.ru"
    YANDEX_PASSWORD = "your_password"

    # Инициализируем наши объекты
    llm_agent = SupplierLLMAgent(
        required_fields=["product_name", "price", "dimensions", "weight", "material"]
    )
    data_manager = SupplierDataManager()
    receiver = YandexEmailReceiver(IMAP_SERVER, YANDEX_EMAIL, YANDEX_PASSWORD)
    sender = YandexEmailSender(SMTP_SERVER, SMTP_PORT, YANDEX_EMAIL, YANDEX_PASSWORD)

    print("Запущен LLM-агент, который реагирует на новые письма на Яндексе...")

    try:
        while True:
            # 1. Проверяем, пришли ли новые письма
            new_emails = receiver.fetch_unseen_emails()

            for msg, from_addr in new_emails:
                # Тема письма (можно использовать в логике)
                subject = msg.get("Subject", "No Subject")

                # Тело письма. Если multipart, то ищем text/plain или text/html
                body_text = ""
                if msg.is_multipart():
                    for part in msg.walk():
                        ctype = part.get_content_type()
                        if ctype in ["text/plain", "text/html"]:
                            try:
                                body_text = part.get_payload(decode=True).decode(part.get_content_charset() or 'utf-8')
                            except:
                                body_text = part.get_payload(decode=True).decode('utf-8', errors='ignore')
                            break
                else:
                    # Не multipart
                    body_text = msg.get_payload(decode=True).decode(msg.get_content_charset() or 'utf-8', errors='ignore')

                # 2. Отправляем текст письма LLM-агенту, чтобы извлечь поля
                parsed = llm_agent.parse_supplier_answer(body_text)

                # 3. Обновляем данные для данного поставщика (по email)
                data_manager.update_data(from_addr, parsed)

                # 4. Проверяем, всё ли заполнено
                if data_manager.is_complete(from_addr, llm_agent.required_fields):
                    print(f"Собраны все данные от поставщика {from_addr}. Сохраняем в Excel.")
                    # Можно сохранить в Excel (или подождать, пока соберём все письма)
                    save_supplier_data_to_excel(data_manager.data, "suppliers_data.xlsx")
                    # Ответим поставщи
